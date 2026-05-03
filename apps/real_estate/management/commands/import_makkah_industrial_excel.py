from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from apps.accounts.models import ProjectMember
from apps.finance.models import Account, Category, Transaction, Transfer
from apps.projects.models import Company, Project
from apps.real_estate.models import IndustrialBuilding, IndustrialCustomerProfile, IndustrialReservationLead, IndustrialUnitRecord


STATUS_MAP = {
    "مؤجر": IndustrialUnitRecord.Status.RENTED,
    "محجوز": IndustrialUnitRecord.Status.RESERVED,
    "غير مؤجر": IndustrialUnitRecord.Status.VACANT,
}


class Command(BaseCommand):
    help = "Import the Makkah North Industrial City operating workbook into editable real-estate records."

    def add_arguments(self, parser):
        parser.add_argument("--file", dest="file_path", default="", help="Path to the source .xlsx workbook.")
        parser.add_argument("--project-name", default="المدينة الصناعية بشمال مكة المكرمة")
        parser.add_argument("--owner", default="", help="Username of the company owner. Defaults to the first superuser/staff/user.")
        parser.add_argument("--clear", action="store_true", help="Clear previous imported industrial records for the project before importing.")
        parser.add_argument("--reset-projects", action="store_true", help="Delete all existing projects before importing this workbook.")

    def handle(self, *args, **options):
        path = self._resolve_file(options["file_path"])
        owner = self._resolve_owner(options["owner"])
        if options["reset_projects"]:
            Transfer.objects.all().delete()
            Transaction.objects.all().delete()
            Account.objects.all().delete()
            Category.objects.all().delete()
            Project.objects.all().delete()
        project = self._get_project(owner, options["project_name"])

        wb = load_workbook(path, data_only=True)
        with transaction.atomic():
            if options["clear"]:
                IndustrialBuilding.objects.filter(project=project).delete()
                IndustrialReservationLead.objects.filter(project=project).delete()
                IndustrialCustomerProfile.objects.filter(project=project).delete()
            building_count, unit_count = self._import_buildings(project, wb)
            lead_count = self._import_leads(project, wb)
            customer_count = self._import_customers(project, wb)

        self.stdout.write(self.style.SUCCESS(
            f"Imported {building_count} buildings, {unit_count} units, {lead_count} leads, {customer_count} customers."
        ))

    def _resolve_file(self, file_path):
        if file_path:
            path = Path(file_path)
        else:
            candidates = sorted(Path.home().joinpath("Downloads").glob("*claude.xlsx"), key=lambda item: item.stat().st_mtime, reverse=True)
            if not candidates:
                raise CommandError("No workbook matching *claude.xlsx was found in Downloads.")
            path = candidates[0]
        if not path.exists():
            raise CommandError(f"Workbook not found: {path}")
        return path

    def _resolve_owner(self, username):
        User = get_user_model()
        if username:
            owner = User.objects.filter(username=username).first()
            if not owner:
                raise CommandError(f"User not found: {username}")
            return owner
        owner = User.objects.filter(is_superuser=True).first() or User.objects.filter(is_staff=True).first() or User.objects.first()
        if not owner:
            raise CommandError("Create a user before importing the workbook.")
        return owner

    def _get_project(self, owner, project_name):
        company, _ = Company.objects.get_or_create(
            owner=owner,
            name="نخبة تسكين العقارية",
            defaults={"country": "SA", "base_currency": "SAR", "description": "إدارة وتشغيل وتأجير"},
        )
        project, _ = Project.objects.get_or_create(
            company=company,
            name=project_name,
            defaults={
                "project_type": Project.ProjectType.REAL_ESTATE,
                "country": "SA",
                "base_currency": "SAR",
                "description": "Makkah North Industrial City",
            },
        )
        ProjectMember.objects.get_or_create(
            user=owner,
            project=project,
            defaults={
                "role": ProjectMember.Role.OWNER,
                "dashboard_access": ProjectMember.DashboardAccess.FULL,
                "can_view_dashboard": True,
                "can_view_financials": True,
                "can_view_raw_excel": True,
                "can_upload_excel": True,
                "can_import_data": True,
                "can_edit_transactions": True,
                "can_export_reports": True,
                "can_manage_users": True,
                "can_view_customers": True,
                "can_view_profit": True,
            },
        )
        return project

    def _import_buildings(self, project, wb):
        unit_count = 0
        building_count = 0
        for ws in wb.worksheets:
            if not ws.title.startswith("مبنى "):
                continue
            title = ws.cell(4, 2).value or ws.title
            activity = self._activity_from_title(str(title))
            declared_units = self._int(ws.cell(7, 2).value)
            building, _ = IndustrialBuilding.objects.update_or_create(
                project=project,
                name=ws.title,
                defaults={"activity": activity, "declared_units": declared_units, "source_sheet": ws.title},
            )
            building.unit_records.all().delete()
            building_count += 1
            for row in range(10, ws.max_row + 1):
                unit_number = ws.cell(row, 3).value
                if unit_number in (None, ""):
                    continue
                record = IndustrialUnitRecord.objects.create(
                    building=building,
                    sequence=self._int(ws.cell(row, 2).value),
                    unit_number=self._clean_number(unit_number),
                    unit_type=self._text(ws.cell(row, 4).value),
                    area=self._decimal(ws.cell(row, 5).value),
                    activity=self._text(ws.cell(row, 6).value),
                    tenant_name=self._text(ws.cell(row, 7).value),
                    phone=self._text(ws.cell(row, 8).value),
                    national_id_or_cr=self._text(ws.cell(row, 9).value),
                    rent_per_meter=self._decimal(ws.cell(row, 10).value),
                    annual_rent=self._decimal(ws.cell(row, 11).value),
                    booking_amount=self._decimal(ws.cell(row, 12).value),
                    paid_amount=self._decimal(ws.cell(row, 13).value),
                    remaining_amount=self._decimal(ws.cell(row, 14).value),
                    contract_start=self._date(ws.cell(row, 15).value),
                    contract_end=self._date(ws.cell(row, 16).value),
                    status=STATUS_MAP.get(self._text(ws.cell(row, 17).value), IndustrialUnitRecord.Status.VACANT),
                    source_row=row,
                    raw_data=self._row_json(ws, row, 2, 17),
                )
                if not record.annual_rent and record.area and record.rent_per_meter:
                    record.annual_rent = record.area * record.rent_per_meter
                    record.remaining_amount = max(record.annual_rent - record.paid_amount, Decimal("0"))
                    record.save(update_fields=["annual_rent", "remaining_amount"])
                unit_count += 1
        return building_count, unit_count

    def _import_leads(self, project, wb):
        ws = next((sheet for sheet in wb.worksheets if "الاستفسارات" in sheet.title), None)
        if not ws:
            return 0
        IndustrialReservationLead.objects.filter(project=project).delete()
        count = 0
        for row in range(6, ws.max_row + 1):
            customer = self._text(ws.cell(row, 4).value)
            if not customer:
                continue
            IndustrialReservationLead.objects.create(
                project=project,
                request_date=self._date(ws.cell(row, 3).value),
                customer_name=customer,
                phone=self._text(ws.cell(row, 5).value),
                building_name=self._text(ws.cell(row, 6).value),
                unit_number=self._clean_number(ws.cell(row, 7).value),
                area=self._text(ws.cell(row, 8).value),
                activity=self._text(ws.cell(row, 9).value),
                rent_value=self._decimal(ws.cell(row, 10).value),
                booking_amount=self._decimal(ws.cell(row, 11).value),
                status=self._text(ws.cell(row, 12).value),
                next_step=self._text(ws.cell(row, 13).value),
                raw_data=self._row_json(ws, row, 2, 13),
            )
            count += 1
        return count

    def _import_customers(self, project, wb):
        ws = next((sheet for sheet in wb.worksheets if "قاعدة بيانات العملاء" in sheet.title), None)
        if not ws:
            return 0
        IndustrialCustomerProfile.objects.filter(project=project).delete()
        count = 0
        for row in range(6, ws.max_row + 1):
            customer = self._text(ws.cell(row, 3).value)
            if not customer:
                continue
            IndustrialCustomerProfile.objects.create(
                project=project,
                customer_name=customer,
                contact_person=self._text(ws.cell(row, 4).value),
                phone=self._text(ws.cell(row, 5).value),
                activity=self._text(ws.cell(row, 6).value),
                budget=self._text(ws.cell(row, 7).value),
                required_units=self._text(ws.cell(row, 8).value),
                building_name=self._text(ws.cell(row, 9).value),
                area=self._text(ws.cell(row, 10).value),
                source=self._text(ws.cell(row, 11).value),
                first_contact=self._text(ws.cell(row, 12).value),
                last_contact=self._text(ws.cell(row, 13).value),
                relationship_status=self._text(ws.cell(row, 14).value),
                next_step=self._text(ws.cell(row, 15).value),
                raw_data=self._row_json(ws, row, 2, 15),
            )
            count += 1
        return count

    def _activity_from_title(self, title):
        if "-" not in title:
            return ""
        return title.split("-", 1)[1].split("(", 1)[0].strip()

    def _text(self, value):
        if value in (None, "—"):
            return ""
        return str(value).strip()

    def _clean_number(self, value):
        if value in (None, ""):
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    def _int(self, value):
        try:
            return int(float(value or 0))
        except (TypeError, ValueError):
            return 0

    def _decimal(self, value):
        if value in (None, "", "—"):
            return Decimal("0")
        try:
            return Decimal(str(value).replace(",", "").replace(" ريال", "").strip())
        except (InvalidOperation, ValueError):
            return Decimal("0")

    def _date(self, value):
        if not value:
            return None
        return value.date() if hasattr(value, "date") else None

    def _row_json(self, ws, row, start_col, end_col):
        headers = [self._text(ws.cell(9 if ws.title.startswith("مبنى ") else 5, col).value) or f"col_{col}" for col in range(start_col, end_col + 1)]
        return {headers[index]: self._json_value(ws.cell(row, col).value) for index, col in enumerate(range(start_col, end_col + 1))}

    def _json_value(self, value):
        if value in (None, "—"):
            return None
        if hasattr(value, "isoformat"):
            return value.isoformat()
        if isinstance(value, Decimal):
            return float(value)
        return value
